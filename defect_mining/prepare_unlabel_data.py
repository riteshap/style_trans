
import json
import os
import openpyxl

model_name = "deepseek-v4-flash"
llm_dict = {"deepseek-v4-flash": "dsv4f", "qwen3.5-flash": "qw35f", "gemini2.5-flash": "gm25f",
                    "llama3.3-70b": "la3370", "nemotron3-ultra": "ne3u"}
model_type = llm_dict[model_name]
task_n = "task7"

wb = openpyxl.Workbook()
ws = wb.active
line_n = 1
for name in os.listdir("./"):
    if f"{task_n}_test" not in name:
        continue
    if f"{model_type}_weakness.json" not in name:
        continue
    level_n = "level-0"
    for l_i in range(6):
        if f"level-{l_i}" in name:
            level_n = f"level-{l_i}"
            break
    file = json.load(open(name, "r", encoding="utf-8"))
    for ele in file:
        src_ind = ele["src_ind"]
        sentence = ele["sentence"]
        weakness_str = ele[model_name]
        weakness_list = json.loads(weakness_str)
        for weakness in weakness_list["weaknesses"]:
            ws[f'A{line_n}'] = src_ind
            ws[f'B{line_n}'] = sentence
            ws[f'C{line_n}'] = weakness
            ws[f'D{line_n}'] = level_n
            line_n += 1

wb.save(f'{task_n}_test_{model_type}_weakness.xlsx')


